import openpyxl
import openpyxl.utils
import openpyxl.comments
import re
from pathlib import Path
import logging
from topline import MONTHS

# imports for openpyxl merge patch
from openpyxl.worksheet import Worksheet
from openpyxl.reader.worksheet import WorkSheetParser
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.merge import MergeCells

logger = logging.getLogger(__name__)

# define excel sheet row references. Updated to excel sheet v7
member_row = 24
header_row = 29
user_row = 30
roi_row = 63
last_roi_row = 67
income_row = 69
last_income_row = 70
expense_row = 72
last_expense_row = 77
account_row = 22
last_account_row = 32


def is_number(s):
    # helper function to test if a value is or can be converted to a number
    try:
        float(s)
        return True
    except ValueError:
        return False


def format_string(s):
    # return list of string or number groups from input string
    # eg: "AA - AUG17" -> ['AA', 'AUG', '17']
    #     "NOV '17 - OCT '18" -> ['NOV', '17', 'OCT', '18']
    s = s.upper()
    string_list = re.findall(r"\d+|[a-z]+", s, re.I)
    string_list = [int(i) if is_number(i) else i.upper() for i in string_list]
    return string_list


def patch_worksheet():
    """This monkeypatches Worksheet.merge_cells to remove cell deletion bug
    https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
    """
    # apply patch 1
    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        cr = CellRange(range_string=range_string, min_col=start_column, min_row=start_row,
                       max_col=end_column, max_row=end_row)
        self.merged_cells.add(cr.coord)
        # self._clean_merge_range(cr)

    Worksheet.merge_cells = merge_cells

    # apply patch 2
    def parse_merge(self, element):
        merged = MergeCells.from_tree(element)
        self.ws.merged_cells.ranges = merged.mergeCell
        # for cr in merged.mergeCell:
        #     self.ws._clean_merge_range(cr)

    WorkSheetParser.parse_merge = parse_merge


patch_worksheet()


class Excel:
    """Class to interface with an Excel workbook
    
    Uses the openpyxl package to read from and write to Excel workbooks.
    
    Attributes:
        filename: The name of the excel file
    
    """
    
    user_ids = None

    def __init__(self, filename):
        self.filename = filename
        self.sheet_names = None
        self.sheet_list = []
        self.header_list = []
        self.summary_sheet = None
        self.workbook = None
        try:
            self.workbook = openpyxl.load_workbook(filename)
            self.get_sheets()
        except FileNotFoundError:
            logger.error("File not found: %s", Path(filename).absolute())
        
        if not Excel.user_ids:
            Excel.user_ids = self.get_user_ids()

        # Force account balances to 0
        logger.info("Clearing all account balances")
        account_range = self.summary_sheet['A' + str(account_row):'C' + str(last_account_row)]
        for r in account_range:
            r[2].value = 0

    def add_transaction(self, transaction):
        """Processes a transaction and writes it to the excel workbook.
        
        Determines the correct sheet, row and column in the workbook that the 
        transaction belongs to and then attempts to write the transaction to 
        the cell depending on the transaction type.
        
        Args:
            transaction: An instance of topline.transaction.Transaction class
        
        Returns:
            True if the transaction is successfully written to the sheet.
            False otherwise.
        """
                
        # determine which sheet to use for current transaction
        sheet = self.get_target_sheet(transaction.month_id, transaction.year % 2000)
        if not sheet:
            logger.warning("Error finding correct sheet for transaction. Month = %s, Year = %s",
                           transaction.month_id, transaction.year % 2000)
            return False
        header = self.get_column_headers(sheet)

        # find correct column for contribution month
        try:
            column = header.index([transaction.month_id, transaction.year % 2000]) + 1
        except ValueError:
            logger.warning("Error finding correct column: Month %s, Year %s",
                           transaction.month_id, transaction.year % 2000)
            return False

        comment = None
        
        # Find correct row in worksheet based on transaction type
        if transaction.type == 'contribution':
            user_offset = [ui for ui, u in enumerate(Excel.user_ids) if u[1] == transaction.username]
            row = user_row + user_offset[0]
        elif transaction.type == 'roi':
            cell = self.find_in_column(transaction.account, sheet, 1, roi_row, last_roi_row)
            if not cell:
                cell = self.find_in_column(None, sheet, 1, roi_row, last_roi_row)
                cell.value = transaction.account
            row = cell.row
        elif transaction.type == 'income':
            cell = self.find_in_column(None, sheet, column, income_row, last_income_row)
            row = cell.row
            comment = transaction.description + ' - ' + transaction.reference
        elif transaction.type == 'expense':
            if 'MONTHLY ACCOUNT FEE' in transaction.description.upper():
                row = expense_row
            else:
                cell = self.find_in_column(None, sheet, column, expense_row + 1, last_expense_row)
                row = cell.row
                comment = transaction.description + ' - ' + transaction.reference
        else:
            return False
        return self.write_to_sheet(sheet, row, column, abs(transaction.amount), add=True, comment=comment)

    def update_account_balances(self, account, balance):
        """Finds account number on the summary sheet and updates its balance"""
        sheet = self.summary_sheet
        cell = self.find_in_column(account, sheet, col=1, start_row=account_row, end_row=last_account_row)
        if cell:
            logger.info("Updating account %s balance: R %.2f", account, balance)
            sheet.cell(row=cell.row, column=3).value = balance
        else:
            logger.info("Adding account %s balance: R %.2f", account, balance)
            new_cell = self.find_in_column(None, sheet, col=1, start_row=account_row, end_row=last_account_row)
            new_cell.value = account
            sheet.cell(row=new_cell.row, column=3).value = balance

    def get_sheets(self):
        """Creates a list of sheet names and a list of month and year indeces"""
        self.sheet_names = self.workbook.sheetnames
        for sheet in self.sheet_names:
            if 'summary' in sheet.lower():
                self.summary_sheet = self.workbook[sheet]
            else:
                # format sheet name
                s = format_string(sheet)
                
                # Iterate through the elements of the sheet name to determine
                # the month and year range that the sheet is for
                name = [i if type(i) is int else (next((j for j, x in MONTHS.items() if i in x), None)) for i in s]
                self.sheet_list.append(name)
        self.sheet_names.remove(self.summary_sheet.title)

    def get_user_ids(self):
        """Returns a list of usernames from the summary sheet."""
        row = 5
        user_ids = []
        while row < self.summary_sheet.max_row:
            user_id = self.summary_sheet.cell(row=row, column=2).value
            if user_id is None:
                break
            else:
                user_ids.append([row - 5, user_id])
                row += 1
        return user_ids

    def get_column_headers(self, sheet):
        """Returns a list of month and year references from column headers."""
        header = sheet[header_row]
        header = [[h.value.month, h.value.year % 2000] if h.is_date else str(h.value) for h in header]
        return header

    def get_target_sheet(self, month, year):
        """Return the sheet to use based on the transaction month and year."""
        sheet_id = [i for i, s in enumerate(self.sheet_list) if
                    (month >= s[0] and year == s[1]) or (month <= s[2] and year == s[3])]
        if not sheet_id:
            return None
        return self.workbook[self.sheet_names[sheet_id[0]]]

    def write_to_sheet(self, sheet, row, col, value, overwrite=False, add=False, comment=None):
        """Writes a value to a cell in a worksheet.
        
        Writes the given value to the cell defined by the row and column
        position in the given sheet, if the cell is not locked. If overwrite
        is True, any value and comment in the cell is overwritten. If add is 
        True, the value and comment are added to the existing value and comment.
        
        Args:
            sheet: a Worksheet instance containing the cell to write to.
            row: the row reference of the cell.
            col: the column reference of the cell.
            value: The value to write.
            overwrite: Optional, Set to True to overwrite existing value and 
                comment in the target cell.
            add: Optional, Set to True to add to the existing value and comment
                in the target cell.
            comment: Optional. The comment to add to the cell.
        
        Returns:
            True if the value and comment were successfully written.
            False otherwise.
        """
        
        target_cell = sheet.cell(row=row, column=col)
        if target_cell.protection.locked:
            logger.warning("[%s] %s%s is locked!", sheet.title, openpyxl.utils.get_column_letter(col), row)
            return False

        new_comment = 'R' + str(value) + ': ' + comment
        cell_val = target_cell.value
        if cell_val == 0 or cell_val == '-' or cell_val is None or overwrite:
            target_cell.value = value
            target_cell.comment = openpyxl.comments.Comment(new_comment, 'Topline') if comment else None
            logger.info("Transaction written to sheet [%s] %s%s.",
                        sheet.title, openpyxl.utils.get_column_letter(col), row)
        elif add:
            logger.info("[%s] %s%s contains data! Adding to existing value!",
                        sheet.title, openpyxl.utils.get_column_letter(col), row)
            target_cell.value = cell_val + value
            
            # Check if there is an existing comment and concatenate it with the
            # new comment.
            old_comment = target_cell.comment.text if target_cell.comment.text else 'R' + str(target_cell.value)
            cell_comment = '\n + '.join(filter(None, (old_comment, new_comment)))
            # cell_comment = cell_comment or (str(cell_val) + ' + ' + str(value))
            target_cell.comment = openpyxl.comments.Comment(cell_comment, 'Topline') if cell_comment else None
        else:
            logger.warning("[%s] %s%s contains data! Transaction not written to sheet!",
                           sheet.title, openpyxl.utils.get_column_letter(col), row)
            return False
        return True

    def close_workbook(self, overwrite=True, filename=None):
        """Save and close the workbook"""
        new_filename = self.filename
        if filename:
            new_filename = filename

        if not overwrite and filename is None:
            logger.warning("No filename provided. Discarding changes")
        else:
            logger.info("Saving workbook to file: %s", new_filename)
            self.workbook.save(new_filename)
        self.workbook.close()

    @staticmethod
    def find_in_column(value, sheet, col, start_row, end_row):
        """Search a column for a value and return the cell if found."""
        if type(col) is int:
            col = openpyxl.utils.get_column_letter(col)
        cell_range = sheet[col + str(start_row):col + str(end_row)]
        cell = [c[0] for c in cell_range if str(value) in str(c[0].value)]
        if not cell:
            return None
        return cell[0]
    
    def set_updating_member(self, username, month, year):
        """Write the member name doing the update to the correct cell"""
        sheet = self.get_target_sheet(month, year % 2000)
        header = self.get_column_headers(sheet)
        # find correct column for current month
        try:
            column = header.index([month, year % 2000]) + 1
        except ValueError:
            logger.warning("Error finding correct column: Month %s, Year %s",
                           month, year % 2000)
            return False
        logger.info("Tracking completed by '%s'. Updating sheet [%s] %s%s.",
                    username, sheet.title, openpyxl.utils.get_column_letter(column), member_row)
        sheet.cell(row=member_row, column=column).value = username
        return True
